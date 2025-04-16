# 优化导入语句
import time
import threading
import configparser
from openpyxl import Workbook
import win32clipboard
import keyboard
from hotkey_manager import HotkeyManager
from window_manager_qt import WindowManagerQt
from model_api import ask_model 
import os
class TextExtractorApp:
    def __init__(self):
        self.word = ""
        self.sentence = ""
        self.data = []
        self.config = self.load_config()  # 新增：读取配置文件
        self.query_loc=0 # 已经查询到的位置
        # 初始化管理器（延迟到QApplication创建后）
        self.window_manager = None
        self.hotkey_manager = HotkeyManager(self)

        
    # 读取配置文件等相关操作
    def load_config(self, config_path='config.ini'):
        """读取配置文件并返回配置对象"""
        import os
        # 检查当前目录
        if not os.path.exists(config_path):
            # 检查exe所在目录
            exe_dir = os.path.dirname(os.path.abspath(__file__))
            config_path = os.path.join(exe_dir, config_path)
        
        config = configparser.ConfigParser()
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config.read_file(f)
            return config
        except Exception as e:
            print(f"读取配置文件失败: {str(e)}")
            self.window_manager._update_log(f"读取配置文件失败: {str(e)}")
            return configparser.ConfigParser()  # 返回空配置对象

    def reset_hotkey(self, hotkey, new_hotkey, callback):
        """重置热键"""
        self.hotkey_manager.change_hotkey(hotkey, new_hotkey, callback)
        self.window_manager._update_log(f"热键已重置为 {new_hotkey}")
        print(f"热键已重置为 {new_hotkey}")  # 调试输出
        if callback==self.add_word :
            self.config.set('hotkeys', "add_data", new_hotkey)  # 更新配置文件
        elif callback == self.capture_word :
            self.config.set('hotkeys', "get_word_hotkey", new_hotkey)  # 更新配置文件
        elif callback == self.capture_sentence :
            self.config.set('hotkeys', "get_sentence_hotkey", new_hotkey)  # 更新配置文件

        with open('config.ini', 'w', encoding='utf-8') as f:  # 保存配置文件
            self.config.write(f)
        self.window_manager.show_hotkeys_and_prompt()

    def run(self):
        """启动程序"""
        # 确保在QApplication上下文中创建窗口
        self.window_manager = WindowManagerQt(self)
        threading.Thread(target=self.hotkey_manager.setup_hotkeys, daemon=True).start()
        self.window_manager.show()

    def get_clipboard_text(self):
        """剪贴板操作封装方法"""
        for _ in range(3):
            try:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()
                
                keyboard.send('ctrl+c')
                time.sleep(0.2)
                
                win32clipboard.OpenClipboard()
                text = win32clipboard.GetClipboardData()
                win32clipboard.CloseClipboard()
                
                if text.strip():
                    return text.strip()
            except Exception as e:
                print(f"剪贴板操作失败: {str(e)}")
                # self.window_manager._update_log(f"剪贴板操作失败: {str(e)}")
                time.sleep(0.1)
        self.window_manager._update_log("错误: 无法获取剪贴板内容")
        return None

    def capture_word(self):
        """捕获单词功能"""
        print(f"检测到{self.config['hotkeys']['get_word_hotkey']}组合键")
        self.window_manager._update_log(f"操作日志: 检测到{self.config['hotkeys']['get_word_hotkey']}组合键")
        time.sleep(0.1)
        self.word = self.get_clipboard_text()
        if self.word:
            self.window_manager.word_input.setText(self.word)
            self.window_manager.log_signal.emit(f"捕获单词: {self.word}")
        else:
            print("未检测到有效文本")
            self.window_manager._update_log("警告: 未检测到有效文本")

    def capture_sentence(self):
        """捕获句子功能"""
        print(f"检测到{self.config['hotkeys']['get_sentence_hotkey']}组合键")
        self.window_manager._update_log(f"操作日志: 检测到{self.config['hotkeys']['get_sentence_hotkey']}组合键")
        time.sleep(0.1)
        self.sentence = self.get_clipboard_text()
        if self.sentence:
            self.window_manager.sentence_input.setText(self.sentence)
            self.window_manager.log_signal.emit(f"捕获句子: {self.sentence}")
        else:
            print("未检测到有效文本")
            self.window_manager._update_log("警告: 未检测到有效句子")

    def load_config(self, config_path='config.ini'):
        """读取配置文件并返回配置对象"""
        import os
        # 检查当前目录
        if not os.path.exists(config_path):
            # 检查exe所在目录
            exe_dir = os.path.dirname(os.path.abspath(__file__))
            config_path = os.path.join(exe_dir, config_path)
        
        config = configparser.ConfigParser()
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config.read_file(f)
            return config
        except Exception as e:
            self.window_manager._update_log(f"错误: 读取配置文件失败: {str(e)}")
            return configparser.ConfigParser()  # 返回空配置对象

    def add_word(self):
        """添加单词功能"""
        print(f"检测到{self.config['hotkeys']['add_data']}组合键")
        self.window_manager._update_log(f"操作日志: 检测到{self.config['hotkeys']['add_data']}组合键")
        self.word = self.window_manager.word_input.text()
        self.sentence = self.window_manager.sentence_input.text()
        if self.word:
            self.data.append({"单词": self.word, "例句": self.sentence})
            self.window_manager._update_log("已保存单词记录")
            self.window_manager.word_input.clear()
            self.window_manager.sentence_input.clear()

    def display_query_result(self, data_item):
        """显示查询结果"""
        keys = list(data_item.keys())
        values = list(data_item.values())
        
        display_info = "查询完成:\n"
        for key, value in zip(keys, values):
            display_info += f"{key}: {value}\n"
        
        self.window_manager._update_log(display_info)

    def model_query(self):
        """模型查词回调函数"""
        def disable_controls():
            """禁用控件"""
            self.window_manager.log_signal.emit("禁用热键和控件")
           #  self.hotkey_manager.remove_hotkeys()
            # 禁用特定按钮
            self.window_manager.save_btn.setEnabled(False)
            self.window_manager.exit_btn.setEnabled(False)
        
        def enable_controls():
            """启用控件"""
           # self.hotkey_manager.setup_hotkeys()
            # 启用特定按钮
            self.window_manager.save_btn.setEnabled(True)
            self.window_manager.exit_btn.setEnabled(True)
            self.window_manager.log_signal.emit("启用热键和控件")
        
        def query_task():
            try:
                disable_controls()
                api_key = self.config["api"]["api_key"]
                self.window_manager._update_log("开始模型查询...")
                for i in range(self.query_loc,len(self.data)):
                    self.window_manager._update_log(f"查询单词: {self.data[i]['单词']}")
                    data = ask_model(api_key, word=self.data[i]["单词"], sentence=self.data[i]["例句"], config=self.config)
                    if data is None:
                        self.window_manager._update_log("api调用失败，检查：1.api_key是否正确，2.使用VPN可能会导致调用失败")
                        break
                    self.data[i] = data
                    self.window_manager.log_signal.emit(f"查询结果: {self.data[i]}")
                    self.query_loc+=1
                self.window_manager._update_log("模型查询完成")
            except Exception as e:
                error_message = f"查询错误: {str(e)}"
                self.window_manager.log_signal.emit(error_message)
            finally:
                enable_controls()
                # 设置查询完成标志
                self._query_complete = True

        if self.data:
            self._query_complete = False  # 新增：查询完成标志
            threading.Thread(target=query_task, daemon=True).start()

    def exit(self):
        """查询按钮点击事件"""
        self.window_manager._update_log("操作日志: 正在保存数据...")
        
        # 启动查询
        self.model_query()
        
        # 等待查询完成
        while not getattr(self, '_query_complete', True):
            time.sleep(0.1)
            self.window_manager.repaint()  # 强制重绘界面保持响应
        
        # 保存数据
        self.save_record()
        self.window_manager._update_log("操作日志: 退出程序")
        
        # 关闭窗口
        self.window_manager.close()



    def on_word_entry_enter(self):
        """单词输入框回车事件处理"""
        word = self.window_manager.word_input.text().strip()
        if word:
            self.word=word
            self.window_manager._update_log(f"已确认单词: {word}")
            self.window_manager.word_input.clearFocus()  # 移除输入框焦点
        else:
            self.window_manager._update_log("错误: 单词不能为空")

    def on_sentence_entry_enter(self):
        """句子输入框回车事件处理"""
        sentence = self.window_manager.sentence_input.text().strip()
        if sentence:
            self.sentence=sentence
            self.window_manager._update_log(f"已确认句子: {sentence}")
            self.window_manager.sentence_input.clearFocus()  # 移除输入框焦点
        else:
            self.window_manager._update_log("错误: 句子不能为空")

    def save_record(self, append_mode=True):
        """保存记录到文件(支持追加模式)"""
        import csv

        
        if not self.data:
            self.window_manager._update_log("警告: 没有数据可保存")
            return

        try:
            # CSV文件保存
            csv_filename = self.config["file"]["output_name"] + ".csv"
            file_exists = os.path.exists(csv_filename)
            
            with open(csv_filename, 'a', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=self.data[0].keys())
                
                if not file_exists or not append_mode:
                    writer.writeheader()
                    
                for item in self.data:
                    writer.writerow(item)
                    
            self.window_manager._update_log(f"✓ 成功保存到 {os.path.abspath(csv_filename)}")

            # # Excel文件保存
            # excel_filename = self.config["file"]["output_name"] + ".xlsx"
            
            # if os.path.exists(excel_filename) and append_mode:
            #     wb = load_workbook(excel_filename)
            #     ws = wb.active
            #     start_row = ws.max_row + 1 if ws.max_row > 1 else 1
            # else:
            #     wb = Workbook()
            #     ws = wb.active
            #     start_row = 1
            #     # 添加表头
            #     header_font = Font(bold=True)
            #     for col, key in enumerate(self.data[0].keys(), 1):
            #         ws.cell(row=1, column=col, value=key).font = header_font

            # # 写入数据
            # for row_idx, item in enumerate(self.data, start=start_row):
            #     for col_idx, value in enumerate(item.values(), 1):
            #         ws.cell(row=row_idx, column=col_idx, value=value)

            # wb.save(excel_filename)
            # self.window_manager._update_log(f"✓ 成功保存到 {os.path.abspath(excel_filename)}")

            # 清空已保存数据
            self.data = []
            self.query_loc = 0
        except Exception as e:
            self.window_manager._update_log(f"保存失败: {str(e)}")



from PySide6.QtWidgets import QApplication
import sys

if __name__ == "__main__":
    qt_app = QApplication(sys.argv)
    app = TextExtractorApp()
    app.run()
    sys.exit(qt_app.exec())
